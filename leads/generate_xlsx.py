#!/usr/bin/env python3
"""Generate XLSX file from CSV lead data for Austrian dental practice prospects."""

import csv
import os

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

def generate_xlsx():
    """Generate the XLSX file from CSV data."""
    if not HAS_OPENPYXL:
        print("openpyxl not available, installing...")
        os.system("pip install openpyxl")
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    
    # === MAIN LEADS SHEET ===
    ws = wb.active
    ws.title = "Dental Practice Leads"
    
    # Read CSV data
    csv_path = os.path.join(os.path.dirname(__file__), "austrian_dental_leads.csv")
    with open(csv_path, 'r', encoding='utf-8') as f:
        reader = csv.reader(f)
        rows = list(reader)
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    high_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    medium_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    low_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Write headers
    for col_idx, header in enumerate(rows[0], 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = thin_border
    
    # Write data rows
    for row_idx, row in enumerate(rows[1:], 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            
            # Color-code by priority
            if len(row) >= 12:
                priority = row[11].strip() if len(row) > 11 else ""
                if priority == "High":
                    cell.fill = high_fill
                elif priority == "Medium":
                    cell.fill = medium_fill
                elif priority == "Low":
                    cell.fill = low_fill
    
    # Set column widths
    col_widths = [25, 22, 18, 35, 30, 20, 7, 25, 50, 12, 12, 10, 60]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    # Freeze panes
    ws.freeze_panes = "A2"
    
    # === TOP 10 HIGHEST PRIORITY SHEET ===
    ws2 = wb.create_sheet("Top 10 Priority")
    
    # Sort by score (descending) and filter High priority
    data_rows = rows[1:]
    high_priority = [r for r in data_rows if len(r) > 11 and r[11].strip() == "High"]
    high_priority.sort(key=lambda x: int(x[6]) if x[6].isdigit() else 0, reverse=True)
    top_10_priority = high_priority[:10]
    
    # Write Top 10 Priority headers
    priority_headers = ["#", "Practice Name", "City", "Score", "Main Weakness", "Website", "Outreach Hook"]
    for col_idx, header in enumerate(priority_headers, 1):
        cell = ws2.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = thin_border
    
    for row_idx, row in enumerate(top_10_priority, 2):
        values = [row_idx - 1, row[0], row[2], row[6], row[7], row[3], row[12] if len(row) > 12 else ""]
        for col_idx, value in enumerate(values, 1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True, vertical='top')
    
    ws2.column_dimensions['A'].width = 5
    ws2.column_dimensions['B'].width = 30
    ws2.column_dimensions['C'].width = 15
    ws2.column_dimensions['D'].width = 8
    ws2.column_dimensions['E'].width = 35
    ws2.column_dimensions['F'].width = 35
    ws2.column_dimensions['G'].width = 60
    ws2.freeze_panes = "A2"
    
    # === TOP 10 EASIEST WINS SHEET ===
    ws3 = wb.create_sheet("Top 10 Easy Wins")
    
    # Easy wins = High score + HTTP-only or weebly/free hosting (clear problems, easy to demonstrate value)
    easy_wins = [r for r in data_rows if len(r) > 11 and (
        'HTTP' in r[7] or 'HTTPS' in r[7] or 'Weebly' in r[7] or 'sta.io' in r[7] or 
        'SSL' in r[8] or 'HTTP' in r[8] or 'free' in r[7].lower()
    )]
    easy_wins.sort(key=lambda x: int(x[6]) if x[6].isdigit() else 0, reverse=True)
    top_10_easy = easy_wins[:10]
    
    # Write Top 10 Easy Wins headers
    for col_idx, header in enumerate(priority_headers, 1):
        cell = ws3.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = thin_border
    
    for row_idx, row in enumerate(top_10_easy, 2):
        values = [row_idx - 1, row[0], row[2], row[6], row[7], row[3], row[12] if len(row) > 12 else ""]
        for col_idx, value in enumerate(values, 1):
            cell = ws3.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True, vertical='top')
    
    ws3.column_dimensions['A'].width = 5
    ws3.column_dimensions['B'].width = 30
    ws3.column_dimensions['C'].width = 15
    ws3.column_dimensions['D'].width = 8
    ws3.column_dimensions['E'].width = 35
    ws3.column_dimensions['F'].width = 35
    ws3.column_dimensions['G'].width = 60
    ws3.freeze_panes = "A2"
    
    # === SUMMARY SHEET ===
    ws4 = wb.create_sheet("Summary Analysis")
    
    ws4.cell(row=1, column=1, value="LEAD GENERATION SUMMARY - Austrian Dental Practices").font = Font(bold=True, size=14)
    ws4.cell(row=2, column=1, value="Prepared for BlueGrid (bluegrid.at)").font = Font(italic=True, size=11)
    ws4.cell(row=3, column=1, value=f"Total Qualified Leads: {len(data_rows)}")
    
    # Count priorities
    high_count = len([r for r in data_rows if len(r) > 11 and r[11].strip() == "High"])
    medium_count = len([r for r in data_rows if len(r) > 11 and r[11].strip() == "Medium"])
    low_count = len([r for r in data_rows if len(r) > 11 and r[11].strip() == "Low"])
    
    ws4.cell(row=4, column=1, value=f"High Priority: {high_count} | Medium Priority: {medium_count} | Low Priority: {low_count}")
    
    ws4.cell(row=6, column=1, value="MOST COMMON WEBSITE PROBLEMS").font = Font(bold=True, size=12)
    problems = [
        "1. No HTTPS / SSL certificate errors (14 practices) - CRITICAL security issue",
        "2. No online booking/Terminanfrage system (40+ practices)",
        "3. Outdated design from 2010-2015 era (20+ practices)",
        "4. Free hosting platforms (Weebly, sta.io) used for professional practice",
        "5. Minimal content - just phone number and address",
        "6. Cookie banners dominating first impression",
        "7. No team/doctor photos visible",
        "8. Poor mobile responsiveness",
        "9. Gmail/non-professional email addresses",
        "10. Long generic text blocks without visual hierarchy",
        "11. No patient testimonials or trust elements",
        "12. Poor SEO structure and meta descriptions"
    ]
    for i, problem in enumerate(problems, 7):
        ws4.cell(row=i, column=1, value=problem)
    
    ws4.column_dimensions['A'].width = 80
    
    # Save
    output_path = os.path.join(os.path.dirname(__file__), "austrian_dental_leads.xlsx")
    wb.save(output_path)
    print(f"✅ XLSX file generated: {output_path}")
    print(f"   Total leads: {len(data_rows)}")
    print(f"   High priority: {high_count}")
    print(f"   Medium priority: {medium_count}")
    print(f"   Low priority: {low_count}")

if __name__ == "__main__":
    generate_xlsx()
